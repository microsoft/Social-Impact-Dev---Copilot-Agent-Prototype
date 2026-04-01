"""Generic FEC CSV formatting and Excel export utilities."""

from __future__ import annotations

import csv
import io

from openpyxl.worksheet.worksheet import Worksheet

from .base import FormCSV
from .constants import HeaderDef


def add_headers_to_csv(parsed: FormCSV) -> str:
    """Add headers to an FEC CSV file.

    Creates a CSV where each record type section has its appropriate headers.

    Args:
        parsed: Parsed form CSV implementing FormCSV protocol.

    Returns:
        CSV string with headers added for each section.
    """
    output = io.StringIO()
    writer = csv.writer(output)

    # Write header record
    hdr_section = parsed.get_header_section()
    if hdr_section:
        writer.writerow(hdr_section.headers)
        for row in hdr_section.rows:
            writer.writerow(row)
        writer.writerow([])

    # Write main sections
    for section in parsed.get_sections():
        writer.writerow(section.headers)
        for row in section.rows:
            writer.writerow(row)
        writer.writerow([])

    return output.getvalue()


def create_xlsx(parsed: FormCSV) -> bytes:
    """Convert parsed FEC CSV to XLSX with multiple sheets.

    Creates an Excel workbook with sheets for each section in the parsed form.

    Args:
        parsed: Parsed form CSV implementing FormCSV protocol.

    Returns:
        XLSX file as bytes.
    """
    try:
        import openpyxl
    except ImportError as e:
        raise ImportError(
            "openpyxl is required for XLSX export. Install with: pip install openpyxl"
        ) from e

    wb = openpyxl.Workbook()
    formatter = FormatService()

    # Remove default sheet, we'll add our own
    wb.remove(wb.active)

    # Create sheets from sections
    for section in parsed.get_sections():
        ws = wb.create_sheet(section.name)
        ws.append(section.headers)
        for row in section.rows:
            ws.append(row)
        formatter.format_sheet(ws, section.columns)

    # Save to bytes
    output = io.BytesIO()
    wb.save(output)
    return output.getvalue()


class FormatService:
    """Service for formatting Excel worksheets."""

    def __init__(
        self,
        *,
        header_color: str = "4472C4",
        header_font_color: str = "FFFFFF",
        min_col_width: int = 12,
        max_col_width: int = 50,
    ) -> None:
        self.header_color = header_color
        self.header_font_color = header_font_color
        self.min_col_width = min_col_width
        self.max_col_width = max_col_width

    def format_sheet(self, ws: Worksheet, columns: list[HeaderDef] | None = None) -> None:
        """Format a worksheet with header styling and auto-width columns."""
        self._format_header_row(ws)
        self._auto_width_columns(ws)

    def _format_header_row(self, ws: Worksheet) -> None:
        """Format the first row as a header with styling and auto-filter."""
        try:
            from openpyxl.styles import Alignment, Font, PatternFill
            from openpyxl.utils import get_column_letter
        except ImportError:
            return

        header_fill = PatternFill(
            start_color=self.header_color,
            end_color=self.header_color,
            fill_type="solid",
        )
        header_font = Font(bold=True, color=self.header_font_color)

        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", wrap_text=True)

        ws.freeze_panes = "A2"

        if ws.max_column > 0 and ws.max_row > 1:
            last_col = get_column_letter(ws.max_column)
            ws.auto_filter.ref = f"A1:{last_col}{ws.max_row}"

    def _auto_width_columns(self, ws: Worksheet) -> None:
        """Auto-size columns based on content."""
        try:
            from openpyxl.utils import get_column_letter
        except ImportError:
            return

        for col_idx, column_cells in enumerate(ws.columns, 1):
            max_length = 0
            column_cells_list = list(column_cells)

            for cell in column_cells_list:
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except Exception:
                    pass

            adjusted_width = max(
                min(max_length + 3, self.max_col_width),
                self.min_col_width,
            )
            ws.column_dimensions[get_column_letter(col_idx)].width = adjusted_width
