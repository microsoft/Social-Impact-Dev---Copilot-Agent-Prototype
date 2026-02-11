"""FEC CSV processing service for adding headers and converting to XLSX."""

from __future__ import annotations

import csv
import io
from dataclasses import dataclass, field

# FEC version 8.x headers with friendly display names
# Source: https://github.com/newsdev/fec-csv-sources

HDR_HEADERS = [
    "Record Type",
    "EF Type",
    "FEC Version",
    "Software Name",
    "Software Version",
    "Name Delimiter",
    "Report ID",
    "Report Number",
    "Comment",
]

# Form 3 (F3/F3N) headers - main filing summary
F3_HEADERS = [
    "Form Type",
    "Committee ID",
    "Committee Name",
    "Change of Address",
    "Street 1",
    "Street 2",
    "City",
    "State",
    "ZIP",
    "Election State",
    "Election District",
    "Report Code",
    "Election Code",
    "Election Date",
    "State of Election",
    "Coverage From Date",
    "Coverage Through Date",
    "Treasurer Last Name",
    "Treasurer First Name",
    "Treasurer Middle Name",
    "Treasurer Prefix",
    "Treasurer Suffix",
    "Date Signed",
    "Total Contributions (No Loans) - Period",
    "Total Contribution Refunds - Period",
    "Net Contributions - Period",
    "Total Operating Expenditures - Period",
    "Total Offset to Operating Expenditures - Period",
    "Net Operating Expenditures - Period",
    "Cash on Hand Close of Period",
    "Debts To",
    "Debts By",
    "Individual Contributions Itemized - Period",
    "Individual Contributions Unitemized - Period",
    "Total Individual Contributions - Period",
    "Political Party Contributions - Period",
    "PAC Contributions - Period",
    "Candidate Contributions - Period",
    "Total Contributions - Period",
    "Transfers From Authorized - Period",
    "Candidate Loans - Period",
    "Other Loans - Period",
    "Total Loans - Period",
    "Offset to Operating Expenditures - Period",
    "Other Receipts - Period",
    "Total Receipts - Period",
    "Operating Expenditures - Period",
    "Transfers to Authorized - Period",
    "Candidate Loan Repayments - Period",
    "Other Loan Repayments - Period",
    "Total Loan Repayments - Period",
    "Refunds to Individuals - Period",
    "Refunds to Party Committees - Period",
    "Refunds to Other Committees - Period",
    "Total Refunds - Period",
    "Other Disbursements - Period",
    "Total Disbursements - Period",
    "Cash on Hand Beginning of Period",
    "Total Receipts This Period",
    "Subtotals",
    "Total Disbursements This Period",
    "Cash on Hand Close",
    "Total Contributions (No Loans) - Cycle",
    "Total Contribution Refunds - Cycle",
    "Net Contributions - Cycle",
    "Total Operating Expenditures - Cycle",
    "Total Offset to Operating Expenditures - Cycle",
    "Net Operating Expenditures - Cycle",
    "Individual Contributions Itemized - Cycle",
    "Individual Contributions Unitemized - Cycle",
    "Total Individual Contributions - Cycle",
    "Political Party Contributions - Cycle",
    "PAC Contributions - Cycle",
    "Candidate Contributions - Cycle",
    "Total Contributions - Cycle",
    "Transfers From Authorized - Cycle",
    "Candidate Loans - Cycle",
    "Other Loans - Cycle",
    "Total Loans - Cycle",
    "Offset to Operating Expenditures - Cycle",
    "Other Receipts - Cycle",
    "Total Receipts - Cycle",
    "Operating Expenditures - Cycle",
    "Transfers to Authorized - Cycle",
    "Candidate Loan Repayments - Cycle",
    "Other Loan Repayments - Cycle",
    "Total Loan Repayments - Cycle",
    "Refunds to Individuals - Cycle",
    "Refunds to Party Committees - Cycle",
    "Refunds to Other Committees - Cycle",
    "Total Refunds - Cycle",
    "Other Disbursements - Cycle",
    "Total Disbursements - Cycle",
]

# Schedule A headers (contributions) - SA11AI, SA11C, SA12, SA14, etc.
SCHEDULE_A_HEADERS = [
    "Form Type",
    "Committee ID",
    "Transaction ID",
    "Back Reference Transaction ID",
    "Back Reference Schedule",
    "Entity Type",
    "Organization Name",
    "Last Name",
    "First Name",
    "Middle Name",
    "Prefix",
    "Suffix",
    "Street 1",
    "Street 2",
    "City",
    "State",
    "ZIP",
    "Election Code",
    "Election Other Description",
    "Date",
    "Amount",
    "Aggregate",
    "Purpose Code",
    "Purpose Description",
    "Employer",
    "Occupation",
    "Donor Committee FEC ID",
    "Donor Committee Name",
    "Donor Candidate FEC ID",
    "Donor Candidate Last Name",
    "Donor Candidate First Name",
    "Donor Candidate Middle Name",
    "Donor Candidate Prefix",
    "Donor Candidate Suffix",
    "Donor Candidate Office",
    "Donor Candidate State",
    "Donor Candidate District",
    "Conduit Name",
    "Conduit Street 1",
    "Conduit Street 2",
    "Conduit City",
    "Conduit State",
    "Conduit ZIP",
    "Memo Code",
    "Memo / Earmark Info",
    "Reference Code",
]

# Schedule B headers (disbursements) - SB17, SB20A, SB21, etc.
SCHEDULE_B_HEADERS = [
    "Form Type",
    "Committee ID",
    "Transaction ID",
    "Back Reference Transaction ID",
    "Back Reference Schedule",
    "Entity Type",
    "Payee Organization Name",
    "Payee Last Name",
    "Payee First Name",
    "Payee Middle Name",
    "Payee Prefix",
    "Payee Suffix",
    "Payee Street 1",
    "Payee Street 2",
    "Payee City",
    "Payee State",
    "Payee ZIP",
    "Election Code",
    "Election Other Description",
    "Date",
    "Amount",
    "Semi-Annual Refunded Bundled Amount",
    "Purpose Code",
    "Purpose Description",
    "Category Code",
    "Beneficiary Committee FEC ID",
    "Beneficiary Committee Name",
    "Beneficiary Candidate FEC ID",
    "Beneficiary Candidate Last Name",
    "Beneficiary Candidate First Name",
    "Beneficiary Candidate Middle Name",
    "Beneficiary Candidate Prefix",
    "Beneficiary Candidate Suffix",
    "Beneficiary Candidate Office",
    "Beneficiary Candidate State",
    "Beneficiary Candidate District",
    "Conduit Name",
    "Conduit Street 1",
    "Conduit Street 2",
    "Conduit City",
    "Conduit State",
    "Conduit ZIP",
    "Memo Code",
    "Memo / Earmark Info",
    "Reference Code",
]

# Mapping of form type prefixes to their headers and friendly names
FORM_TYPE_CONFIG: dict[str, tuple[list[str], str]] = {
    "HDR": (HDR_HEADERS, "Header"),
    "F3": (F3_HEADERS, "Summary"),
    "SA": (SCHEDULE_A_HEADERS, "Contributions"),
    "SB": (SCHEDULE_B_HEADERS, "Disbursements"),
}


@dataclass
class ParsedFECFile:
    """Parsed FEC filing with records grouped by type."""

    version: str
    header: list[str]
    summary: list[str]
    contributions: list[list[str]] = field(default_factory=list)
    disbursements: list[list[str]] = field(default_factory=list)
    other: list[list[str]] = field(default_factory=list)
    all_rows: list[list[str]] = field(default_factory=list)


def _get_headers_for_form_type(form_type: str) -> list[str]:
    """Get the appropriate headers for a form type."""
    form_type_upper = form_type.upper().strip('"')

    # Check exact match first
    if form_type_upper in FORM_TYPE_CONFIG:
        return FORM_TYPE_CONFIG[form_type_upper][0]

    # Check prefix match (e.g., SA11AI -> SA, SB17 -> SB, F3N -> F3)
    for prefix, (headers, _) in FORM_TYPE_CONFIG.items():
        if form_type_upper.startswith(prefix):
            return headers

    return []


def _pad_row(row: list[str], headers: list[str]) -> list[str]:
    """Pad a row to match the number of headers."""
    if len(row) < len(headers):
        return row + [""] * (len(headers) - len(row))
    return row[: len(headers)]


def parse_fec_csv(content: bytes | str) -> ParsedFECFile:
    """Parse an FEC CSV file and categorize records by type.

    Args:
        content: Raw CSV content as bytes or string.

    Returns:
        ParsedFECFile with records grouped by type.
    """
    if isinstance(content, bytes):
        content = content.decode("utf-8", errors="replace")

    reader = csv.reader(io.StringIO(content))
    rows = list(reader)

    result = ParsedFECFile(
        version="",
        header=[],
        summary=[],
    )

    for row in rows:
        if not row:
            continue

        result.all_rows.append(row)
        form_type = row[0].strip('"').upper()

        if form_type == "HDR":
            result.header = row
            # Extract version (typically in column 3)
            if len(row) > 2:
                result.version = row[2].strip('"')
        elif form_type.startswith("F3"):
            result.summary = row
        elif form_type.startswith("SA"):
            result.contributions.append(row)
        elif form_type.startswith("SB"):
            result.disbursements.append(row)
        else:
            result.other.append(row)

    return result


def add_headers_to_csv(content: bytes | str) -> str:
    """Add headers to an FEC CSV file.

    Creates a CSV where each record type section has its appropriate headers.

    Args:
        content: Raw FEC CSV content.

    Returns:
        CSV string with headers added for each section.
    """
    parsed = parse_fec_csv(content)
    output = io.StringIO()
    writer = csv.writer(output)

    # Write header record with its headers
    if parsed.header:
        writer.writerow(HDR_HEADERS[: len(parsed.header)])
        writer.writerow(parsed.header)
        writer.writerow([])  # Empty row separator

    # Write summary record with its headers
    if parsed.summary:
        headers = _get_headers_for_form_type(parsed.summary[0])
        if headers:
            writer.writerow(headers[: len(parsed.summary)])
        writer.writerow(parsed.summary)
        writer.writerow([])

    # Write contributions with headers
    if parsed.contributions:
        writer.writerow(SCHEDULE_A_HEADERS)
        for row in parsed.contributions:
            writer.writerow(_pad_row(row, SCHEDULE_A_HEADERS))
        writer.writerow([])

    # Write disbursements with headers
    if parsed.disbursements:
        writer.writerow(SCHEDULE_B_HEADERS)
        for row in parsed.disbursements:
            writer.writerow(_pad_row(row, SCHEDULE_B_HEADERS))
        writer.writerow([])

    # Write other records
    if parsed.other:
        writer.writerow(["form_type", "data..."])
        for row in parsed.other:
            writer.writerow(row)

    return output.getvalue()


def create_xlsx(content: bytes | str) -> bytes:
    """Convert FEC CSV to XLSX with multiple sheets.

    Creates an Excel workbook with:
    - "Raw" sheet: Exact copy of original CSV data
    - "Summary" sheet: Filing summary (F3 record)
    - "Contributions" sheet: All Schedule A records with headers
    - "Disbursements" sheet: All Schedule B records with headers

    Args:
        content: Raw FEC CSV content.

    Returns:
        XLSX file as bytes.
    """
    try:
        import openpyxl
    except ImportError as e:
        raise ImportError(
            "openpyxl is required for XLSX export. Install with: pip install openpyxl"
        ) from e

    parsed = parse_fec_csv(content)
    wb = openpyxl.Workbook()

    # Raw sheet - exact copy of CSV
    ws_raw = wb.active
    ws_raw.title = "Raw"
    for row in parsed.all_rows:
        ws_raw.append(row)

    # Summary sheet
    if parsed.summary:
        ws_summary = wb.create_sheet("Summary")
        headers = _get_headers_for_form_type(parsed.summary[0])
        if headers:
            ws_summary.append(headers[: len(parsed.summary)])
        ws_summary.append(parsed.summary)
        _format_header_row(ws_summary)

    # Contributions sheet
    if parsed.contributions:
        ws_contrib = wb.create_sheet("Contributions")
        ws_contrib.append(SCHEDULE_A_HEADERS)
        for row in parsed.contributions:
            ws_contrib.append(_pad_row(row, SCHEDULE_A_HEADERS))
        _format_header_row(ws_contrib)
        _auto_width_columns(ws_contrib)

    # Disbursements sheet
    if parsed.disbursements:
        ws_disb = wb.create_sheet("Disbursements")
        ws_disb.append(SCHEDULE_B_HEADERS)
        for row in parsed.disbursements:
            ws_disb.append(_pad_row(row, SCHEDULE_B_HEADERS))
        _format_header_row(ws_disb)
        _auto_width_columns(ws_disb)

    # Save to bytes
    output = io.BytesIO()
    wb.save(output)
    return output.getvalue()


def _format_header_row(ws) -> None:
    """Format the first row as a header (bold)."""
    try:
        from openpyxl.styles import Font
    except ImportError:
        return

    for cell in ws[1]:
        cell.font = Font(bold=True)
    ws.freeze_panes = "A2"


def _auto_width_columns(ws, max_width: int = 50) -> None:
    """Auto-size columns based on content."""
    try:
        from openpyxl.utils import get_column_letter
    except ImportError:
        return

    for col_idx, column_cells in enumerate(ws.columns, 1):
        max_length = 0
        for cell in column_cells:
            try:
                cell_length = len(str(cell.value)) if cell.value else 0
                max_length = max(max_length, cell_length)
            except Exception:
                pass
        adjusted_width = min(max_length + 2, max_width)
        ws.column_dimensions[get_column_letter(col_idx)].width = adjusted_width
